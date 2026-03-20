#!/usr/bin/env python
"""
xlsx_diff.py - Compare two .xlsx files and output a git-diff-style result.

Usage:
    uv run python tools/xlsx_diff.py old.xlsx new.xlsx [--sheet <name>] [--no-color] [--no-inline]

Exit codes:
    0 - No differences found
    1 - Differences found
"""

import sys
import argparse
import difflib
import openpyxl

# ---------------------------------------------------------------------------
# Optional: diff-match-patch for character-level inline diff
# ---------------------------------------------------------------------------
try:
    import diff_match_patch as dmp_module
    _DMP_AVAILABLE = True
except ImportError:
    _DMP_AVAILABLE = False

# ---------------------------------------------------------------------------
# ANSI color helpers
# ---------------------------------------------------------------------------
_RESET       = "\033[0m"
_RED_BG_BOLD = "\033[1;41m"   # bold + red background  (deletions)
_GRN_BG_BOLD = "\033[1;42m"   # bold + green background (insertions)
_RED_FG      = "\033[31m"     # red foreground for "-" lines
_GRN_FG      = "\033[32m"     # green foreground for "+" lines
_CYAN_FG     = "\033[36m"     # cyan foreground for hunk headers


def _supports_color(no_color: bool) -> bool:
    """Return True when ANSI color should be used."""
    if no_color:
        return False
    return sys.stdout.isatty()


# ---------------------------------------------------------------------------
# xlsx reading
# ---------------------------------------------------------------------------

def _load_sheet_rows(path: str, sheet_name: str | None) -> dict[str, list[str]]:
    """
    Load an xlsx file and return {sheet_name: [row_text, ...]} for each sheet.
    If sheet_name is given, only load that sheet.
    Row text is tab-joined cell values; fully-empty rows are omitted.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        print(f"Error opening {path}: {exc}", file=sys.stderr)
        sys.exit(2)

    sheets = [sheet_name] if sheet_name else wb.sheetnames
    result: dict[str, list[str]] = {}
    for name in sheets:
        if name not in wb.sheetnames:
            print(f"Sheet '{name}' not found in {path}", file=sys.stderr)
            sys.exit(2)
        ws = wb[name]
        rows: list[str] = []
        for row in ws.iter_rows():
            cells = [str(c.value) if c.value is not None else "" for c in row]
            if any(v.strip() for v in cells):
                rows.append("\t".join(cells))
        result[name] = rows
    return result


# ---------------------------------------------------------------------------
# Inline (character-level) diff helpers
# ---------------------------------------------------------------------------

def _inline_diff_color(old_line: str, new_line: str) -> tuple[str, str]:
    """
    Use diff-match-patch to highlight character-level changes.
    Returns (colored_old, colored_new) with ANSI codes.
    """
    dmp = dmp_module.diff_match_patch()
    diffs = dmp.diff_main(old_line, new_line)
    dmp.diff_cleanupSemantic(diffs)

    old_parts: list[str] = []
    new_parts: list[str] = []
    for op, text in diffs:
        if op == dmp_module.diff_match_patch.DIFF_EQUAL:
            old_parts.append(text)
            new_parts.append(text)
        elif op == dmp_module.diff_match_patch.DIFF_DELETE:
            old_parts.append(f"{_RED_BG_BOLD}{text}{_RESET}")
        elif op == dmp_module.diff_match_patch.DIFF_INSERT:
            new_parts.append(f"{_GRN_BG_BOLD}{text}{_RESET}")

    return "".join(old_parts), "".join(new_parts)


def _inline_diff_plain(old_line: str, new_line: str) -> tuple[str, str]:
    """
    Use diff-match-patch to highlight character-level changes.
    Returns (marked_old, marked_new) with [-text-] / [+text+] markers.
    """
    dmp = dmp_module.diff_match_patch()
    diffs = dmp.diff_main(old_line, new_line)
    dmp.diff_cleanupSemantic(diffs)

    old_parts: list[str] = []
    new_parts: list[str] = []
    for op, text in diffs:
        if op == dmp_module.diff_match_patch.DIFF_EQUAL:
            old_parts.append(text)
            new_parts.append(text)
        elif op == dmp_module.diff_match_patch.DIFF_DELETE:
            old_parts.append(f"[-{text}-]")
        elif op == dmp_module.diff_match_patch.DIFF_INSERT:
            new_parts.append(f"[+{text}+]")

    return "".join(old_parts), "".join(new_parts)


# ---------------------------------------------------------------------------
# Core diff logic
# ---------------------------------------------------------------------------

def _diff_sheet(
    sheet_name: str,
    old_rows: list[str],
    new_rows: list[str],
    use_color: bool,
    use_inline: bool,
    context: int = 2,
) -> list[str]:
    """
    Produce a unified-diff-style list of output lines for one sheet.
    Returns an empty list when there are no differences.
    """
    output: list[str] = []
    matcher = difflib.SequenceMatcher(None, old_rows, new_rows, autojunk=False)
    opcodes = matcher.get_opcodes()

    def _hunk_header(old_start: int, old_end: int, new_start: int, new_end: int) -> str:
        hdr = f"@@ -{old_start + 1},{old_end - old_start} +{new_start + 1},{new_end - new_start} @@"
        return f"{_CYAN_FG}{hdr}{_RESET}" if use_color else hdr

    def _fmt_del(line: str) -> str:
        return (f"{_RED_FG}-{line}{_RESET}" if use_color else f"-{line}")

    def _fmt_add(line: str) -> str:
        return (f"{_GRN_FG}+{line}{_RESET}" if use_color else f"+{line}")

    def _fmt_ctx(line: str) -> str:
        return f" {line}"

    # Build hunks with context
    # Collect groups of opcodes that are "near" each other (within 2*context lines)
    groups = list(matcher.get_grouped_opcodes(context))

    if not groups:
        return []  # no differences

    sheet_header = f"=== Sheet: {sheet_name} ==="
    output.append(sheet_header)

    for group in groups:
        # Determine hunk span
        first = group[0]
        last  = group[-1]
        output.append(_hunk_header(first[1], last[2], first[3], last[4]))

        for tag, i1, i2, j1, j2 in group:
            if tag == "equal":
                for line in old_rows[i1:i2]:
                    output.append(_fmt_ctx(line))

            elif tag == "replace":
                old_block = old_rows[i1:i2]
                new_block = new_rows[j1:j2]

                if use_inline and _DMP_AVAILABLE and len(old_block) == len(new_block):
                    # Pair-wise character diff
                    for old_line, new_line in zip(old_block, new_block):
                        if use_color:
                            o, n = _inline_diff_color(old_line, new_line)
                        else:
                            o, n = _inline_diff_plain(old_line, new_line)
                        output.append(_fmt_del(o))
                        output.append(_fmt_add(n))
                else:
                    for line in old_block:
                        output.append(_fmt_del(line))
                    for line in new_block:
                        output.append(_fmt_add(line))

            elif tag == "delete":
                for line in old_rows[i1:i2]:
                    output.append(_fmt_del(line))

            elif tag == "insert":
                for line in new_rows[j1:j2]:
                    output.append(_fmt_add(line))

    return output


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Compare two .xlsx files and show diff."
    )
    parser.add_argument("old", help="Old xlsx file")
    parser.add_argument("new", help="New xlsx file")
    parser.add_argument("--sheet", default=None, help="Only diff a specific sheet")
    parser.add_argument("--no-color", action="store_true", help="Disable ANSI color output")
    parser.add_argument("--no-inline", action="store_true", help="Disable character-level inline diff")
    args = parser.parse_args()

    use_color  = _supports_color(args.no_color)
    use_inline = not args.no_inline

    if use_inline and not _DMP_AVAILABLE:
        # Silently degrade; no error
        use_inline = False

    old_data = _load_sheet_rows(args.old, args.sheet)
    new_data = _load_sheet_rows(args.new, args.sheet)

    # Union of sheet names, preserving order
    all_sheets: list[str] = []
    seen: set[str] = set()
    for name in list(old_data.keys()) + list(new_data.keys()):
        if name not in seen:
            all_sheets.append(name)
            seen.add(name)

    has_diff = False
    for sheet_name in all_sheets:
        old_rows = old_data.get(sheet_name, [])
        new_rows = new_data.get(sheet_name, [])
        diff_lines = _diff_sheet(sheet_name, old_rows, new_rows, use_color, use_inline)
        if diff_lines:
            has_diff = True
            for line in diff_lines:
                print(line)

    sys.exit(1 if has_diff else 0)


if __name__ == "__main__":
    main()

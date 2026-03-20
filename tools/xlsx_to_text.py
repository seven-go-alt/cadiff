#!/usr/bin/env python
"""
xlsx_to_text.py - Convert an .xlsx file to plain text (tab-separated).

Used as a git textconv driver:
    git config diff.xlsx.textconv "uv run python tools/xlsx_to_text.py"
    git config diff.xlsx.binary true

Usage:
    uv run python tools/xlsx_to_text.py <file.xlsx>
"""

import sys
import openpyxl


def xlsx_to_text(path: str) -> None:
    """Read an xlsx file and print all sheets as tab-separated text."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"Error opening {path}: {e}", file=sys.stderr)
        sys.exit(1)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows():
            values = [str(cell.value) if cell.value is not None else "" for cell in row]
            # Skip entirely empty rows
            if not any(v.strip() for v in values):
                continue
            print("\t".join(values))


def main() -> None:
    if len(sys.argv) != 2:
        print(f"Usage: {sys.argv[0]} <file.xlsx>", file=sys.stderr)
        sys.exit(1)

    xlsx_to_text(sys.argv[1])


if __name__ == "__main__":
    main()

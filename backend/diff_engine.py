"""
diff_engine.py — Core diff logic that returns structured JSON.

Reuses the row/cell reading approach from tools/xlsx_diff.py but outputs
a rich data structure instead of ANSI text, suitable for JSON serialization
and frontend rendering.
"""

from __future__ import annotations

import difflib
import io
from typing import Any

import openpyxl

try:
    import diff_match_patch as dmp_module
    _DMP_AVAILABLE = True
except ImportError:
    _DMP_AVAILABLE = False


# ---------------------------------------------------------------------------
# xlsx reading
# ---------------------------------------------------------------------------

def _load_sheet_data(wb: openpyxl.Workbook) -> dict[str, list[list[str]]]:
    """Return {sheet_name: [[cell, …], …]} for all non-empty rows."""
    result: dict[str, list[list[str]]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows: list[list[str]] = []
        for row in ws.iter_rows():
            cells = [str(c.value) if c.value is not None else "" for c in row]
            if any(v.strip() for v in cells):
                rows.append(cells)
        result[name] = rows
    return result


def _row_to_str(cells: list[str]) -> str:
    return "\t".join(cells)


# ---------------------------------------------------------------------------
# Character-level diff (diff-match-patch)
# ---------------------------------------------------------------------------

def _inline_diff(old_text: str, new_text: str) -> list[dict[str, str]]:
    """Character-level diff via diff-match-patch."""
    if not _DMP_AVAILABLE:
        return [{"text": old_text, "op": "delete"}, {"text": new_text, "op": "insert"}]

    dmp = dmp_module.diff_match_patch()
    diffs = dmp.diff_main(old_text, new_text)
    dmp.diff_cleanupSemantic(diffs)

    _OP = {
        dmp_module.diff_match_patch.DIFF_EQUAL: "equal",
        dmp_module.diff_match_patch.DIFF_DELETE: "delete",
        dmp_module.diff_match_patch.DIFF_INSERT: "insert",
    }
    return [{"text": text, "op": _OP[op]} for op, text in diffs]


def _build_inline_cells(
    old_cells: list[str], new_cells: list[str]
) -> list[list[dict[str, str]]]:
    """Per-column character-level diff segments."""
    max_len = max(len(old_cells), len(new_cells))
    oc = old_cells + [""] * (max_len - len(old_cells))
    nc = new_cells + [""] * (max_len - len(new_cells))

    result: list[list[dict[str, str]]] = []
    for o, n in zip(oc, nc):
        if o == n:
            result.append([{"text": o, "op": "equal"}])
        else:
            result.append(_inline_diff(o, n))
    return result


# ---------------------------------------------------------------------------
# Metadata helpers
# ---------------------------------------------------------------------------

def _extract_headers(
    old_rows: list[list[str]], new_rows: list[list[str]]
) -> list[str] | None:
    if old_rows:
        return old_rows[0]
    if new_rows:
        return new_rows[0]
    return None


def _count_stats(hunks: list[dict[str, Any]]) -> dict[str, int]:
    added = deleted = modified = 0
    for hunk in hunks:
        for row in hunk["rows"]:
            t = row["type"]
            if t == "insert":
                added += 1
            elif t == "delete":
                deleted += 1
            elif t == "replace":
                modified += 1
    return {"added": added, "deleted": deleted, "modified": modified}


def _has_usable_key_column(
    old_rows: list[list[str]], new_rows: list[list[str]]
) -> bool:
    """
    True when the first column looks like a unique ID within each file
    (>=70% unique values per file, skipping the assumed header row).
    """
    def _uniq(rows: list[list[str]]) -> float:
        data = rows[1:]
        if not data:
            return 0.0
        cols = [r[0] for r in data if r]
        return len(set(cols)) / len(cols) if cols else 0.0

    return _uniq(old_rows) >= 0.7 and _uniq(new_rows) >= 0.7


# ---------------------------------------------------------------------------
# Context-windowing on a flat row list
# ---------------------------------------------------------------------------

def _build_hunks_from_flat(
    flat: list[dict[str, Any]], context: int
) -> list[dict[str, Any]]:
    """
    Apply context windowing to a flat ordered list of diff rows, returning
    hunk dicts {old_start, old_count, new_start, new_count, rows}.
    """
    n = len(flat)
    changed = [i for i, r in enumerate(flat) if r["type"] != "equal"]
    if not changed:
        return []

    include: set[int] = set()
    for ci in changed:
        for j in range(max(0, ci - context), min(n, ci + context + 1)):
            include.add(j)

    sorted_idx = sorted(include)
    spans: list[tuple[int, int]] = []
    start = prev = sorted_idx[0]
    for idx in sorted_idx[1:]:
        if idx > prev + 1:
            spans.append((start, prev + 1))
            start = idx
        prev = idx
    spans.append((start, prev + 1))

    hunks = []
    for a, b in spans:
        rows = flat[a:b]
        old_nos = [r["old_row_no"] for r in rows if r.get("old_row_no") is not None]
        new_nos = [r["new_row_no"] for r in rows if r.get("new_row_no") is not None]
        hunks.append({
            "old_start": old_nos[0] if old_nos else 0,
            "old_count": len(old_nos),
            "new_start": new_nos[0] if new_nos else 0,
            "new_count": len(new_nos),
            "rows": rows,
        })
    return hunks


# ---------------------------------------------------------------------------
# Key-based differ (first column = row ID)
# ---------------------------------------------------------------------------

def _diff_by_key(
    old_rows: list[list[str]], new_rows: list[list[str]], context: int
) -> list[dict[str, Any]]:
    """
    Run SequenceMatcher on first-column values (keys).

    equal opcode  → same key; check content to emit 'equal' or 'replace'
    replace/delete → old rows with unmatched keys → 'delete'
    replace/insert → new rows with unmatched keys → 'insert'
    """
    old_keys = [r[0] if r else "" for r in old_rows]
    new_keys = [r[0] if r else "" for r in new_rows]

    matcher = difflib.SequenceMatcher(None, old_keys, new_keys, autojunk=False)

    flat: list[dict[str, Any]] = []
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            for oi, ni in zip(range(i1, i2), range(j1, j2)):
                old_r, new_r = old_rows[oi], new_rows[ni]
                if old_r == new_r:
                    flat.append({
                        "type": "equal",
                        "old_row_no": oi + 1, "new_row_no": ni + 1,
                        "old_cells": old_r, "new_cells": new_r, "inline": None,
                    })
                else:
                    # Same ID, different content → modified row
                    flat.append({
                        "type": "replace",
                        "old_row_no": oi + 1, "new_row_no": ni + 1,
                        "old_cells": old_r, "new_cells": new_r,
                        "inline": _build_inline_cells(old_r, new_r),
                    })
        elif tag in ("replace", "delete"):
            # Keys only in old → deleted rows
            for oi in range(i1, i2):
                flat.append({
                    "type": "delete",
                    "old_row_no": oi + 1, "new_row_no": None,
                    "old_cells": old_rows[oi], "new_cells": None, "inline": None,
                })
            if tag == "replace":
                # Keys only in new → inserted rows
                for ni in range(j1, j2):
                    flat.append({
                        "type": "insert",
                        "old_row_no": None, "new_row_no": ni + 1,
                        "old_cells": None, "new_cells": new_rows[ni], "inline": None,
                    })
        elif tag == "insert":
            for ni in range(j1, j2):
                flat.append({
                    "type": "insert",
                    "old_row_no": None, "new_row_no": ni + 1,
                    "old_cells": None, "new_cells": new_rows[ni], "inline": None,
                })

    return _build_hunks_from_flat(flat, context)


# ---------------------------------------------------------------------------
# Sequence-based differ (full row strings)
# ---------------------------------------------------------------------------

def _diff_by_sequence(
    old_rows: list[list[str]], new_rows: list[list[str]], context: int
) -> list[dict[str, Any]]:
    """Fallback: SequenceMatcher on tab-joined row strings."""
    old_strs = [_row_to_str(r) for r in old_rows]
    new_strs = [_row_to_str(r) for r in new_rows]
    matcher = difflib.SequenceMatcher(None, old_strs, new_strs, autojunk=False)

    hunks = []
    for group in matcher.get_grouped_opcodes(context):
        first, last = group[0], group[-1]
        hunk_rows: list[dict[str, Any]] = []

        for tag, i1, i2, j1, j2 in group:
            if tag == "equal":
                for oi, ni in zip(range(i1, i2), range(j1, j2)):
                    hunk_rows.append({
                        "type": "equal",
                        "old_row_no": oi + 1, "new_row_no": ni + 1,
                        "old_cells": old_rows[oi], "new_cells": new_rows[ni],
                        "inline": None,
                    })
            elif tag == "replace":
                ob, nb = old_rows[i1:i2], new_rows[j1:j2]
                min_len = min(len(ob), len(nb))
                # Pair as many rows as possible sequentially
                for k in range(min_len):
                    o, n = ob[k], nb[k]
                    hunk_rows.append({
                        "type": "replace",
                        "old_row_no": i1 + k + 1, "new_row_no": j1 + k + 1,
                        "old_cells": o, "new_cells": n,
                        "inline": _build_inline_cells(o, n),
                    })
                # Leftover old rows are deleted
                for k in range(min_len, len(ob)):
                    hunk_rows.append({
                        "type": "delete",
                        "old_row_no": i1 + k + 1, "new_row_no": None,
                        "old_cells": ob[k], "new_cells": None, "inline": None,
                    })
                # Leftover new rows are inserted
                for k in range(min_len, len(nb)):
                    hunk_rows.append({
                        "type": "insert",
                        "old_row_no": None, "new_row_no": j1 + k + 1,
                        "old_cells": None, "new_cells": nb[k], "inline": None,
                    })
            elif tag == "delete":
                for k, o in enumerate(old_rows[i1:i2]):
                    hunk_rows.append({
                        "type": "delete",
                        "old_row_no": i1 + k + 1, "new_row_no": None,
                        "old_cells": o, "new_cells": None, "inline": None,
                    })
            elif tag == "insert":
                for k, n in enumerate(new_rows[j1:j2]):
                    hunk_rows.append({
                        "type": "insert",
                        "old_row_no": None, "new_row_no": j1 + k + 1,
                        "old_cells": None, "new_cells": n, "inline": None,
                    })

        hunks.append({
            "old_start": first[1] + 1, "old_count": last[2] - first[1],
            "new_start": first[3] + 1, "new_count": last[4] - first[3],
            "rows": hunk_rows,
        })
    return hunks


# ---------------------------------------------------------------------------
# Sheet and workbook entry points
# ---------------------------------------------------------------------------

def _diff_sheets(
    old_data: dict[str, list[list[str]]],
    new_data: dict[str, list[list[str]]],
    context: int = 2,
) -> list[dict[str, Any]]:
    seen: set[str] = set()
    all_sheets: list[str] = []
    for name in list(old_data) + list(new_data):
        if name not in seen:
            all_sheets.append(name)
            seen.add(name)

    sheets: list[dict[str, Any]] = []

    for sheet_name in all_sheets:
        old_rows = old_data.get(sheet_name, [])
        new_rows = new_data.get(sheet_name, [])

        headers = _extract_headers(old_rows, new_rows)
        use_key = _has_usable_key_column(old_rows, new_rows)

        if use_key:
            hunks = _diff_by_key(old_rows, new_rows, context)
        else:
            hunks = _diff_by_sequence(old_rows, new_rows, context)

        has_diff = any(r["type"] != "equal" for h in hunks for r in h["rows"])
        stats = _count_stats(hunks)

        sheets.append({
            "name": sheet_name,
            "has_diff": has_diff,
            "headers": headers,
            "stats": stats,
            "hunks": hunks,
        })

    return sheets


def diff_workbooks(old_bytes: bytes, new_bytes: bytes) -> dict[str, Any]:
    """Compare two xlsx files given as raw bytes. Returns structured JSON."""
    old_wb = openpyxl.load_workbook(io.BytesIO(old_bytes), data_only=True)
    new_wb = openpyxl.load_workbook(io.BytesIO(new_bytes), data_only=True)

    old_data = _load_sheet_data(old_wb)
    new_data = _load_sheet_data(new_wb)

    sheets = _diff_sheets(old_data, new_data)
    has_diff = any(s["has_diff"] for s in sheets)

    return {"has_diff": has_diff, "sheets": sheets}
